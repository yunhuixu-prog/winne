# -*- coding: utf-8 -*-
"""Render Dayu requirement details to xlsx + markdown table."""

import argparse
import errno
import json
import os
import re
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional


def _auto_install_enabled() -> bool:
    value = os.environ.get("DAYU_AUTO_INSTALL_DEPS", "1").strip().lower()
    return value not in ("0", "false", "no", "off")


def _deps_dir() -> str:
    override = os.environ.get("DAYU_DEPS_DIR")
    if override:
        return override
    home = os.path.expanduser("~")
    return os.path.join(home, ".agents", "skills", ".pydeps")


def _ensure_deps_path() -> str:
    deps_dir = _deps_dir()
    if os.path.isdir(deps_dir) and deps_dir not in sys.path:
        sys.path.insert(0, deps_dir)
    return deps_dir


def _ensure_dependency(import_name: str, pip_name: Optional[str] = None) -> None:
    pip_pkg = pip_name or import_name
    deps_dir = _ensure_deps_path()
    try:
        __import__(import_name)
        return
    except ModuleNotFoundError:
        if not _auto_install_enabled():
            raise RuntimeError(
                f"Missing dependency '{pip_pkg}'. Set DAYU_AUTO_INSTALL_DEPS=1 "
                f"or install it manually (python3 -m pip install --target {deps_dir} {pip_pkg})."
            )

        sys.stderr.write(f"Dependency '{pip_pkg}' missing. Installing via pip...\n")
        try:
            os.makedirs(deps_dir, exist_ok=True)
            subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "pip",
                    "install",
                    "--target",
                    deps_dir,
                    pip_pkg,
                ],
                check=True,
                stdout=sys.stderr,
                stderr=sys.stderr,
            )
        except Exception as exc:
            raise RuntimeError(
                f"Failed to auto-install '{pip_pkg}'. "
                "Please grant permission for pip installs or run: "
                f"python3 -m pip install --target {deps_dir} {pip_pkg}"
            ) from exc
        _ensure_deps_path()
        __import__(import_name)


_ensure_dependency("openpyxl")
from openpyxl import Workbook
from openpyxl.styles import Font


EXPORT_COLUMNS = [
    "需求类型",
    "事件来源",
    "事件类型",
    "*事件id",
    "*事件名称",
    "参数",
    "参数名称",
    "参数类型",
    "参数值类型",
    "参数口径",
    "参数值",
    "参数值名称",
    "参数值口径",
    "*统计口径",
    "备注说明",
    "事件分组",
    "标签",
]

PLATFORM_NAME_MAP = {1: "iOS", 2: "Android", 3: "Harmony"}
PARAM_TYPE_MAP = {1: "普通参数", 2: "私有参数"}
VALUE_TYPE_MAP = {1: "string", 2: "number", 3: "bool", 4: "array", 5: "object"}
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
FALLBACK_OUTPUT_DIR = Path("/tmp/dayu/output")
SKILL_DIR_ENV_KEYS = (
    "SKILL_DIR",
    "CODEX_SKILL_DIR",
    "CURSOR_SKILL_DIR",
    "CLAUDE_SKILL_DIR",
    "GOOGLE_SKILL_DIR",
)


def _safe(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", " ").strip()


def _as_int(value: Any) -> Optional[int]:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _extract_requirements(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, dict) and isinstance(data.get("requirements"), list):
        requirements = data.get("requirements") or []
        if not requirements:
            raise ValueError("requirements is empty.")
        return [item for item in requirements if isinstance(item, dict)]
    if isinstance(data, list):
        if not data:
            raise ValueError("input list is empty.")
        if not isinstance(data[0], dict):
            raise ValueError("input list element must be object.")
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        return [data]
    raise ValueError("Unsupported input JSON structure.")


def _pick_requirement_name(requirement: Dict[str, Any]) -> str:
    for key in ("name", "requirementName", "title"):
        value = _safe(requirement.get(key))
        if value:
            return value
    requirement_id = _safe(requirement.get("id"))
    if requirement_id:
        return f"需求_{requirement_id}"
    return "未命名需求"


def _pick_share_link(requirement: Dict[str, Any]) -> str:
    for key in ("shareLink", "requirementShareLink", "link"):
        value = _safe(requirement.get(key))
        if value:
            return value
    return ""


def _safe_filename(text: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", text.strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = cleaned.strip("._")
    return cleaned or "未命名需求"


def _unique_prefix(base: str, used: set[str]) -> str:
    candidate = base
    index = 2
    while candidate in used:
        candidate = f"{base}_{index}"
        index += 1
    used.add(candidate)
    return candidate


def flatten_rows(requirement: Dict[str, Any]) -> List[List[str]]:
    rows: List[List[str]] = []
    events = requirement.get("events") or []
    for event in events:
        event_source = _safe(event.get("eventSource"))
        event_type = _safe(event.get("eventType"))
        event_id = _safe(event.get("event"))
        event_name = _safe(event.get("eventName"))
        event_info = _safe(event.get("info"))

        for req_type in event.get("requirementTypes") or []:
            req_name = _safe(req_type.get("requirementTypeName"))
            remark = _safe(req_type.get("remark"))
            params = req_type.get("params") or [{}]
            for p in params:
                param_type = _safe(PARAM_TYPE_MAP.get(_as_int(p.get("paramType")), p.get("paramType")))
                value_type = _safe(VALUE_TYPE_MAP.get(_as_int(p.get("valueType")), p.get("valueType")))
                values = p.get("values") or [{}]
                for val in values:
                    rows.append(
                        [
                            req_name,
                            event_source,
                            event_type,
                            event_id,
                            event_name,
                            _safe(p.get("param")),
                            _safe(p.get("paramName")),
                            param_type,
                            value_type,
                            _safe(p.get("info")),
                            _safe(val.get("val")),
                            _safe(val.get("valName")),
                            _safe(val.get("info")),
                            event_info,
                            remark,
                            "",
                            "",
                        ]
                    )
    return rows


def _md_escape(text: str) -> str:
    return text.replace("|", "\\|")


def to_markdown(rows: List[List[str]]) -> str:
    head = "| " + " | ".join(EXPORT_COLUMNS) + " |"
    sep = "|" + "|".join(["---"] * len(EXPORT_COLUMNS)) + "|"
    body = []
    for row in rows:
        body.append("| " + " | ".join(_md_escape(_safe(v)) for v in row) + " |")
    return "\n".join([head, sep] + body)


def render_markdown(rows: List[List[str]]) -> str:
    return to_markdown(rows)


def _group_ranges(rows: List[List[str]], key_indices: List[int]) -> List[tuple[int, int]]:
    ranges: List[tuple[int, int]] = []
    i = 0
    while i < len(rows):
        j = i + 1
        while j < len(rows) and all(rows[j][k] == rows[i][k] for k in key_indices):
            j += 1
        ranges.append((i, j - 1))
        i = j
    return ranges


def _apply_xlsx_merges(ws: Any, rows: List[List[str]], data_start_row: int) -> None:
    if not rows:
        return
    # xlsx detail columns are 1-based:
    # 2=事件来源, 3=事件类型, 4=*事件id, 5=*事件名称, 14=*统计口径
    event_merge_cols = [2, 3, 4, 5, 14]
    # 6=参数, 7=参数名称, 8=参数类型, 9=参数值类型, 10=参数口径, 15=备注说明
    param_merge_cols = [6, 7, 8, 9, 10, 15]

    # keys on flattened row list (0-based)
    event_key_indices = [3]  # *事件id
    param_key_indices = [3, 0, 5]  # *事件id + 需求类型 + 参数

    for start_idx, end_idx in _group_ranges(rows, event_key_indices):
        if end_idx <= start_idx:
            continue
        top = data_start_row + start_idx
        bottom = data_start_row + end_idx
        for col in event_merge_cols:
            ws.merge_cells(start_row=top, start_column=col, end_row=bottom, end_column=col)

    for start_idx, end_idx in _group_ranges(rows, param_key_indices):
        if end_idx <= start_idx:
            continue
        top = data_start_row + start_idx
        bottom = data_start_row + end_idx
        for col in param_merge_cols:
            ws.merge_cells(start_row=top, start_column=col, end_row=bottom, end_column=col)


def export_xlsx(requirement: Dict[str, Any], output_file: Path, mode: str) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "需求明细"

    header_font = Font(bold=True)
    # Row 1: detail header; row 2+: data rows
    ws.append(EXPORT_COLUMNS)
    rows = flatten_rows(requirement)
    for row in rows:
        ws.append(row)
    if mode == "merged":
        _apply_xlsx_merges(ws, rows, data_start_row=2)

    for cell in ws[1]:
        cell.font = header_font

    wb.save(str(output_file))
    return len(rows)


def _is_permission_error(exc: OSError) -> bool:
    return exc.errno in {errno.EACCES, errno.EPERM, errno.EROFS}


def _is_dir_writable(path: Path) -> bool:
    if not os.access(path, os.W_OK):
        return False
    probe = path / ".dayu_write_probe"
    try:
        probe.write_text("ok", encoding="utf-8")
        probe.unlink()
        return True
    except OSError:
        return False


def _prepare_output_dir(custom_output_dir: Optional[str]) -> tuple[Path, Optional[Dict[str, Any]]]:
    if custom_output_dir:
        output_dir = Path(custom_output_dir).expanduser().resolve()
        output_dir.mkdir(parents=True, exist_ok=True)
        if not _is_dir_writable(output_dir):
            raise PermissionError(f"Output directory is not writable: {output_dir}")
        return output_dir, None

    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        if not _is_dir_writable(OUTPUT_DIR):
            raise PermissionError(errno.EACCES, "Directory is not writable", str(OUTPUT_DIR))
        return OUTPUT_DIR, None
    except OSError as exc:
        if not _is_permission_error(exc):
            raise
        FALLBACK_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        return (
            FALLBACK_OUTPUT_DIR,
            {
                "from": str(OUTPUT_DIR),
                "to": str(FALLBACK_OUTPUT_DIR),
                "reason": "default output directory is not writable",
            },
        )


def _resolve_skill_root(cli_skill_root: Optional[str]) -> Path:
    script_path = Path(__file__).resolve()
    inferred = script_path.parent.parent
    if cli_skill_root:
        candidate = Path(cli_skill_root).expanduser().resolve()
        skill_md = candidate / "SKILL.md"
        expected_script = candidate / "scripts" / script_path.name
        if skill_md.is_file() and expected_script == script_path:
            return candidate
        raise ValueError(
            f"Invalid --skill-root: {cli_skill_root}. "
            "Expected a skill root that contains SKILL.md and this script under scripts/."
        )

    candidates: list[Path] = []
    for key in SKILL_DIR_ENV_KEYS:
        value = os.environ.get(key, "").strip()
        if value:
            path = Path(value).expanduser().resolve()
            if path not in candidates:
                candidates.append(path)
    candidates.append(inferred)

    for candidate in candidates:
        skill_md = candidate / "SKILL.md"
        expected_script = candidate / "scripts" / script_path.name
        if skill_md.is_file() and expected_script == script_path:
            return candidate

    raise ValueError(
        "Unable to resolve current skill root. Set --skill-root or SKILL_DIR; global file search is disabled."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Dayu requirement xlsx + markdown table.")
    parser.add_argument("--input", default=None, help="Requirement JSON file")
    parser.add_argument(
        "--input-json",
        default=None,
        help="Requirement JSON string. Preferred when avoiding temporary files.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Output directory for generated xlsx/md files. Defaults to skill-level output directory.",
    )
    parser.add_argument(
        "--skill-root",
        default=None,
        help="Lock skill root directory explicitly. Resolution order: --skill-root, SKILL_DIR/agent envs, script location.",
    )
    parser.add_argument(
        "--mode",
        choices=["expanded", "merged"],
        default="expanded",
        help="xlsx mode: expanded(default) or merged. Markdown always expanded.",
    )
    args = parser.parse_args()

    if bool(args.input) == bool(args.input_json):
        raise ValueError("Provide exactly one of --input or --input-json.")
    _resolve_skill_root(args.skill_root)

    if args.input_json:
        data = json.loads(args.input_json)
    else:
        data = json.loads(Path(args.input).read_text(encoding="utf-8"))

    requirements = _extract_requirements(data)
    if not requirements:
        raise ValueError("No valid requirement object found in input JSON.")

    output_dir, output_dir_fallback = _prepare_output_dir(args.output_dir)
    used_prefixes: set[str] = set()
    exports: List[Dict[str, Any]] = []
    for req in requirements:
        req_name = _pick_requirement_name(req)
        req_id = _safe(req.get("id"))
        share_link = _pick_share_link(req)
        base = _safe_filename(req_name)
        if req_id:
            base = f"{base}_{_safe_filename(req_id)}"
        prefix = _unique_prefix(base, used_prefixes)

        output_xlsx = output_dir / f"{prefix}.xlsx"
        output_md = output_dir / f"{prefix}.md"

        rows = flatten_rows(req)
        row_count = export_xlsx(req, output_xlsx, mode=args.mode)
        output_md.write_text(render_markdown(rows), encoding="utf-8")
        exports.append(
            {
                "requirementName": req_name,
                "requirementId": req_id,
                "shareLink": share_link,
                "rowCount": row_count,
                "xlsx": str(output_xlsx),
                "markdown": str(output_md),
            }
        )

    export_summaries = [
        {
            "mdFileName": Path(item["markdown"]).name,
            "xlsxFileName": Path(item["xlsx"]).name,
            "shareLink": item.get("shareLink", ""),
            "requirementName": item.get("requirementName", ""),
        }
        for item in exports
    ]

    result: Dict[str, Any] = {
        "outputDir": str(output_dir),
        "count": len(exports),
        "mode": args.mode,
        "exports": exports,
        "exportSummaries": export_summaries,
        "notice": f"All exported files are saved in: {output_dir}",
    }
    if output_dir_fallback:
        result["outputDirFallback"] = output_dir_fallback
        result["notice"] = (
            f"Default output directory is not writable. Files were saved in fallback path: {output_dir}"
        )
    if len(exports) == 1:
        one = exports[0]
        result.update(
            {
                "rowCount": one["rowCount"],
                "xlsx": one["xlsx"],
                "markdown": one["markdown"],
                "requirementId": one["requirementId"],
                "requirementName": one["requirementName"],
                "shareLink": one["shareLink"],
            }
        )

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
