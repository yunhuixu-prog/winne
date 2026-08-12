#!/usr/bin/env python3
"""将最新续费率_真实和预估.csv 按列并排导出 Excel（每 4 列一国一端，预估期绿色底纹）。"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

Key = Tuple[str, str, str]  # country, os, period_type
Cell = Tuple[float, str]  # rate, source

SEGMENTS = [
    ("美国", "android", "美国Android"),
    ("美国", "ios", "美国iOS"),
    ("英国", "android", "英国Android"),
    ("英国", "ios", "英国iOS"),
    ("巴西", "android", "巴西Android"),
    ("巴西", "ios", "巴西iOS"),
    ("墨西哥", "android", "墨西哥Android"),
    ("墨西哥", "ios", "墨西哥iOS"),
    ("西班牙", "android", "西班牙Android"),
    ("西班牙", "ios", "西班牙iOS"),
    ("加拿大", "android", "加拿大Android"),
    ("加拿大", "ios", "加拿大iOS"),
    ("澳大利亚", "android", "澳大利亚Android"),
    ("澳大利亚", "ios", "澳大利亚iOS"),
    ("其他", "android", "其他Android"),
    ("其他", "ios", "其他iOS"),
]
PERIOD_TYPES = ("月", "年", "周")
RATE_HEADERS = ("月", "年", "周")
COLS_PER_SEGMENT = 3  # 月、年、周；续费次数仅 A 列
MAX_K = 100
DATA_START_ROW = 3  # 第 1 行标题，第 2 行列名，第 3 行起为 1～100 期

FILL_REAL = PatternFill(fill_type="solid", fgColor="FFFFFF")
FILL_FORECAST = PatternFill(fill_type="solid", fgColor="C6EFCE")
FONT_HEADER = Font(bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def load_data(path: Path) -> Dict[Key, Dict[int, Cell]]:
    data: Dict[Key, Dict[int, Cell]] = {}
    with path.open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            key: Key = (
                r["country_name"].strip(),
                r["os_type"].strip(),
                r["period_type"].strip(),
            )
            k = int(r["renewal_k"])
            rate = float(r["renewal_rate"]) if (r.get("renewal_rate") or "").strip() else 0.0
            src = (r.get("renewal_rate_source") or "").strip()
            data.setdefault(key, {})[k] = (rate, src)
    return data


def set_rate_cell(ws, row: int, col: int, series: Dict[int, Cell], k: int) -> None:
    cell = ws.cell(row, col)
    if k in series:
        rate, src = series[k]
        cell.value = rate
        cell.number_format = "0.00%"
        cell.fill = FILL_FORECAST if src == "预估" else FILL_REAL
    else:
        cell.value = ""
        cell.fill = FILL_REAL
    cell.alignment = ALIGN_CENTER


def write_workbook(data: Dict[Key, Dict[int, Cell]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "最新续费率"

    # A 列：续费次数（全表共用）
    ws.cell(2, 1, "续费次数").font = FONT_HEADER
    ws.cell(2, 1).alignment = ALIGN_CENTER
    for k in range(1, MAX_K + 1):
        row = DATA_START_ROW + k - 1
        ws.cell(row, 1, k).alignment = ALIGN_CENTER

    for seg_idx, (country, os_type, title) in enumerate(SEGMENTS):
        base_col = 2 + seg_idx * COLS_PER_SEGMENT  # B-D, E-G, ...

        # 第 1 行：单元标题
        ws.cell(1, base_col, title).font = FONT_HEADER

        # 第 2 行：月、年、周
        for offset, hdr in enumerate(RATE_HEADERS):
            c = ws.cell(2, base_col + offset, hdr)
            c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER

        series_by_pt = {
            pt: data.get((country, os_type, pt), {}) for pt in PERIOD_TYPES
        }

        # 第 3～102 行：1～100 期（续费次数在 A 列，此处只写月/年/周）
        for k in range(1, MAX_K + 1):
            row = DATA_START_ROW + k - 1
            set_rate_cell(ws, row, base_col, series_by_pt["月"], k)
            set_rate_cell(ws, row, base_col + 1, series_by_pt["年"], k)
            set_rate_cell(ws, row, base_col + 2, series_by_pt["周"], k)

    total_cols = 1 + len(SEGMENTS) * COLS_PER_SEGMENT
    ws.column_dimensions["A"].width = 10
    for col in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.freeze_panes = "B3"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    skill_root = Path(__file__).resolve().parent.parent
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--input",
        type=Path,
        default=skill_root / "out" / "20260519" / "最新续费率_真实和预估.csv",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=skill_root / "out" / "20260519" / "最新续费率_展示.xlsx",
    )
    args = ap.parse_args()
    data = load_data(args.input)
    write_workbook(data, args.out)
    print(f"wrote -> {args.out}")


if __name__ == "__main__":
    main()
