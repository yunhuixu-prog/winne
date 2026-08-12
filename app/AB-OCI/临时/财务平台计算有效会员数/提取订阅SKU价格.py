"""
预估 Excel → 订阅 SKU 连续包价格宽表

复用指令:
  从预估 Excel 各「指标_{国家}_{端口}」sheet 提取「新增订阅SKU价格」「续费订阅SKU价格」
  中的连续包月/季/年/周，生成宽表 CSV：行=月份，列按
  西班牙、其他、墨西哥、美国、巴西、加拿大、英国、澳大利亚 × iOS、Android ×
  新增月、新增季、新增年、新增周、续费月、续费季、续费年、续费周 排列，
  列名 {国家}_{平台}_{指标}，只输出 各国家端口_订阅SKU连续包价格_宽表.csv

用法:
  python3 提取订阅SKU价格.py [输入xlsx路径]
"""

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

INPUT_XLSX = '5月预估-最新版.xlsx'
OUTPUT_WIDE_CSV = '各国家端口_订阅SKU连续包价格_宽表.csv'

TARGET_SKUS = ['连续包月', '连续包季', '连续包年', '连续包周']
PRICE_SECTIONS = ['新增订阅SKU价格', '续费订阅SKU价格']

COUNTRY_ORDER = ['西班牙', '其他', '墨西哥', '美国', '巴西', '加拿大', '英国', '澳大利亚']
PORT_ORDER = ['iOS', 'Android']
METRIC_ORDER = ['新增月', '新增季', '新增年', '新增周', '续费月', '续费季', '续费年', '续费周']

METRIC_MAP = {
    ('新增订阅SKU价格', '连续包月'): '新增月',
    ('新增订阅SKU价格', '连续包季'): '新增季',
    ('新增订阅SKU价格', '连续包年'): '新增年',
    ('新增订阅SKU价格', '连续包周'): '新增周',
    ('续费订阅SKU价格', '连续包月'): '续费月',
    ('续费订阅SKU价格', '连续包季'): '续费季',
    ('续费订阅SKU价格', '连续包年'): '续费年',
    ('续费订阅SKU价格', '连续包周'): '续费周',
}


def parse_sheet_name(name: str):
    m = re.match(r'^指标_(.+?)_(.+)$', name)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def find_section_columns(ws, section_name: str, target_skus: list[str]):
    start_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == section_name:
            start_col = c
            break
    if start_col is None:
        return {}

    mapping = {}
    c = start_col
    while c <= ws.max_column:
        header = ws.cell(1, c).value
        sku = ws.cell(2, c).value
        if header and header != section_name and header not in (None, ''):
            if header in PRICE_SECTIONS:
                break
            if header.endswith('退款率') or header.endswith('占比') or '分成' in str(header):
                break
            break
        if sku in target_skus:
            mapping[sku] = c
        c += 1
        if len(mapping) == len(target_skus):
            break
    return mapping


def extract_sheet(ws, country: str, port: str):
    rows = []
    section_cols = {s: find_section_columns(ws, s, TARGET_SKUS) for s in PRICE_SECTIONS}

    for r in range(3, ws.max_row + 1):
        month = ws.cell(r, 1).value
        if month is None or month == '':
            continue
        try:
            month = int(month)
        except (TypeError, ValueError):
            continue

        for section, col_map in section_cols.items():
            if not col_map:
                continue
            for sku, col in col_map.items():
                val = ws.cell(r, col).value
                if val is None or val == '':
                    continue
                rows.append({
                    '国家': country,
                    '端口': port,
                    '月份': month,
                    '价格类型': section,
                    'SKU类型': sku,
                    '价格': float(val),
                })
    return rows


def port_to_platform(port: str) -> str:
    if port.endswith('iOS'):
        return 'iOS'
    if port.endswith('Android'):
        return 'Android'
    return port


def build_wide_table(df: pd.DataFrame) -> pd.DataFrame:
    tmp = df.copy()
    tmp['平台'] = tmp['端口'].map(port_to_platform)
    tmp['指标'] = tmp.apply(
        lambda r: METRIC_MAP.get((r['价格类型'], r['SKU类型'])), axis=1
    )
    tmp = tmp.dropna(subset=['指标'])

    wide = tmp.pivot_table(
        index='月份',
        columns=['国家', '平台', '指标'],
        values='价格',
        aggfunc='first',
    )

    full_keys = [
        (country, platform, metric)
        for country in COUNTRY_ORDER
        for platform in PORT_ORDER
        for metric in METRIC_ORDER
    ]
    for key in full_keys:
        if key not in wide.columns:
            wide[key] = pd.NA
    wide = wide[full_keys]
    wide.columns = [f'{c}_{p}_{m}' for c, p, m in full_keys]
    return wide.reset_index().sort_values('月份')


def main():
    base_dir = Path(__file__).parent
    xlsx_name = sys.argv[1] if len(sys.argv) > 1 else INPUT_XLSX
    xlsx_path = Path(xlsx_name) if Path(xlsx_name).is_absolute() else base_dir / xlsx_name

    print(f'读取: {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    all_rows = []
    for sn in sorted(s for s in wb.sheetnames if s.startswith('指标_') and s != '指标_业务汇总'):
        country, port = parse_sheet_name(sn)
        if country:
            all_rows.extend(extract_sheet(wb[sn], country, port))

    df = pd.DataFrame(all_rows)
    if df.empty:
        print('未提取到数据')
        return

    wide_df = build_wide_table(df)
    out_path = base_dir / OUTPUT_WIDE_CSV
    wide_df.to_csv(out_path, index=False, encoding='utf-8-sig')
    print(f'已输出: {out_path} ({len(wide_df)} 行 x {len(wide_df.columns)} 列)')


if __name__ == '__main__':
    main()
