"""
历史未订阅老用户试用分析数据处理

输入: 试用用户分析.csv
输出:
  - 老用户_各层级来源占比及付费转化率.csv  (表1)
  - 老用户_三级来源功能使用比例.csv        (表2，含列均值行)
  - 老用户_三级来源功能使用比例.xlsx       (表2，高于列均值标红)
  - 老用户_使用其他功能后付费和留存转化率.csv  (表3)
  - 老用户_功能叠加推荐.csv
  - 老用户_功能组合推荐.csv
"""

import pandas as pd
import numpy as np
from itertools import combinations
from pathlib import Path

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------
INPUT_CSV = '试用用户分析.csv'
USER_TYPE = '历史未订阅老用户'
NON_FUNC_SOURCES = {'无三级', 'Onboarding', 'Update First Launch'}
MIN_FUNC_COL_RATE = 0.01       # 表2 功能列保留阈值
MIN_PAIR_SAMPLES = 100         # 组合推荐最小样本
MIN_LIFT_REC = 0.02            # 叠加推荐最小 lift
MIN_PEN_REC = 0.10             # 叠加推荐最小使用占比

OUTPUT_TABLE1 = '老用户_各层级来源占比及付费转化率.csv'
OUTPUT_TABLE2 = '老用户_三级来源功能使用比例.csv'
OUTPUT_TABLE2_XLSX = '老用户_三级来源功能使用比例.xlsx'

TABLE2_META_COLS = {
    '三级来源', '用户数', '占比', '付费人数', '付费转化率', '留存率',
    '使用该三级功能比例', '除该三级功能还使用其他功能比例',
    '仅使用来源功能占比', '仅使用来源功能付费率', '仅使用来源功能留存率',
    '还使用其他功能占比', '还使用其他功能付费率', '还使用其他功能留存率',
    '试用其他来源功能比仅使用来源功能留存率提升',
}
OUTPUT_TABLE3 = '老用户_使用其他功能后付费和留存转化率.csv'
OUTPUT_ADD_REC = '老用户_功能叠加推荐.csv'
OUTPUT_PAIR_REC = '老用户_功能组合推荐.csv'


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def parse_funcs(val):
    if pd.isna(val) or str(val).strip() in ('', '\\N', 'nan'):
        return set()
    return {f.strip() for f in str(val).split(',') if f.strip() and f.strip() != '\\N'}


def get_third_source(row):
    """三级来源分组规则（表1/表2/表3 共用）"""
    fs, ss = row['first_source'], row['second_source']
    if fs == 'Edit' and ss in ('Retouch', 'Edit', 'Material'):
        third = '' if pd.isna(row['third_source']) else str(row['third_source']).strip()
        return third if third not in ('', 'nan', 'None') else '无三级'
    if fs == 'Else' and ss in ('Onboarding', 'Update First Launch'):
        return ss
    return '无三级'


def fmt_pct(series):
    return series.apply(lambda x: f'{x:.2%}' if pd.notna(x) else '-')


def load_old_users(csv_path=INPUT_CSV):
    df = pd.read_csv(csv_path)
    old = df[df['user_type'].str.contains(USER_TYPE, na=False)].copy()
    old['third_src'] = old.apply(get_third_source, axis=1)
    old['use_funcs'] = old['use_func_list'].apply(parse_funcs)
    return old


# ---------------------------------------------------------------------------
# 表1: 各层级来源占比及付费转化率
# ---------------------------------------------------------------------------
def build_table1(old):
    total = len(old)
    rows = []
    for level, col in [
        ('一级来源', 'first_source'),
        ('二级来源', 'second_source'),
        ('三级来源', 'third_src'),
    ]:
        tmp = old.copy()
        if col == 'third_src':
            tmp['source_name'] = tmp['third_src']
        else:
            tmp['source_name'] = tmp[col].fillna('未知').astype(str).str.strip()
            tmp.loc[tmp['source_name'].isin(['', 'nan', 'None']), 'source_name'] = '未知'
        grp = (
            tmp.groupby('source_name')
            .agg(用户数=('gid', 'count'), 付费人数=('is_paid', 'sum'))
            .reset_index()
        )
        grp['占比'] = grp['用户数'] / total
        grp['付费转化率'] = grp['付费人数'] / grp['用户数']
        grp['来源层级'] = level
        rows.append(grp.rename(columns={'source_name': '来源名称'}))
    out = pd.concat(rows, ignore_index=True)
    return out[['来源层级', '来源名称', '用户数', '占比', '付费人数', '付费转化率']]


# ---------------------------------------------------------------------------
# 表2: 三级来源功能使用比例
# ---------------------------------------------------------------------------
def build_table2(old):
    total = len(old)
    all_funcs = sorted(set().union(*old['use_funcs']))
    rows = []

    for src, grp in old.groupby('third_src'):
        n = len(grp)
        is_func = src not in NON_FUNC_SOURCES
        base_paid = grp['is_paid'].mean()

        use_src_cnt = use_other_cnt = only_src_cnt = 0
        only_src_paid = other_paid = 0
        only_src_retained = other_retained = 0
        func_cnt = {f: 0 for f in all_funcs}

        for funcs, paid, retained in zip(
            grp['use_funcs'], grp['is_paid'], grp['is_retention_1'],
        ):
            for f in funcs:
                func_cnt[f] += 1
            if is_func and src in funcs:
                use_src_cnt += 1
                if funcs == {src}:
                    only_src_cnt += 1
                    only_src_paid += paid
                    only_src_retained += retained
                elif any(f != src for f in funcs):
                    use_other_cnt += 1
                    other_paid += paid
                    other_retained += retained

        base_retention = grp['is_retention_1'].mean()
        only_src_ret_rate = only_src_retained / only_src_cnt if only_src_cnt else np.nan
        other_ret_rate = other_retained / use_other_cnt if use_other_cnt else np.nan

        row = {
            '三级来源': src,
            '用户数': n,
            '占比': n / total,
            '付费人数': int(grp['is_paid'].sum()),
            '付费转化率': base_paid,
            '留存率': base_retention,
        }
        if is_func:
            row['使用该三级功能比例'] = use_src_cnt / n
            row['除该三级功能还使用其他功能比例'] = use_other_cnt / n
            row['仅使用来源功能占比'] = only_src_cnt / n
            row['仅使用来源功能付费率'] = only_src_paid / only_src_cnt if only_src_cnt else np.nan
            row['仅使用来源功能留存率'] = only_src_ret_rate
            row['还使用其他功能占比'] = use_other_cnt / n
            row['还使用其他功能付费率'] = other_paid / use_other_cnt if use_other_cnt else np.nan
            row['还使用其他功能留存率'] = other_ret_rate
            row['试用其他来源功能比仅使用来源功能留存率提升'] = (
                other_ret_rate - only_src_ret_rate
                if pd.notna(other_ret_rate) and pd.notna(only_src_ret_rate) else np.nan
            )
        else:
            for k in (
                '使用该三级功能比例', '除该三级功能还使用其他功能比例',
                '仅使用来源功能占比', '仅使用来源功能付费率', '仅使用来源功能留存率',
                '还使用其他功能占比', '还使用其他功能付费率', '还使用其他功能留存率',
                '试用其他来源功能比仅使用来源功能留存率提升',
            ):
                row[k] = np.nan
        for f in all_funcs:
            row[f] = func_cnt[f] / n
        rows.append(row)

    out = pd.DataFrame(rows).sort_values('用户数', ascending=False)
    keep_funcs = [
        f for f in all_funcs
        if out[f].max() >= MIN_FUNC_COL_RATE or f in out['三级来源'].values
    ]
    base_cols = [
        '三级来源', '用户数', '占比', '付费人数', '付费转化率', '留存率',
        '使用该三级功能比例', '除该三级功能还使用其他功能比例',
        '仅使用来源功能占比', '仅使用来源功能付费率', '仅使用来源功能留存率',
        '还使用其他功能占比', '还使用其他功能付费率', '还使用其他功能留存率',
        '试用其他来源功能比仅使用来源功能留存率提升',
    ]
    return out[base_cols + keep_funcs]


# ---------------------------------------------------------------------------
# 表3: 三级来源下使用其他各功能后的付费转化率
# ---------------------------------------------------------------------------
def build_table3(old):
    source_order = old['third_src'].value_counts().index.tolist()
    order_map = {s: i for i, s in enumerate(source_order)}
    rows = []

    for src, grp in old.groupby('third_src'):
        n = len(grp)
        base_paid = grp['is_paid'].mean()
        base_retention = grp['is_retention_1'].mean()
        is_func = src not in NON_FUNC_SOURCES
        all_f = set().union(*grp['use_funcs'])

        for f in sorted(all_f):
            if is_func and f == src:
                continue
            sub = grp[grp['use_funcs'].apply(lambda x, ff=f: ff in x)]
            cnt = len(sub)
            if cnt == 0:
                continue
            paid = sub['is_paid'].sum()
            retained = sub['is_retention_1'].sum()
            rate = paid / cnt
            ret_rate = retained / cnt
            rows.append({
                '三级来源': src,
                '其他功能': f,
                '使用人数': cnt,
                '使用占比': cnt / n,
                '付费人数': int(paid),
                '付费转化率': rate,
                '入口基准付费率': base_paid,
                'lift': rate - base_paid,
                '留存率': ret_rate,
                '入口基准留存率': base_retention,
                '留存率lift': ret_rate - base_retention,
            })

    out = pd.DataFrame(rows)
    out['_src_ord'] = out['三级来源'].map(order_map)
    out = out.sort_values(['_src_ord', '使用人数'], ascending=[True, False])
    return out.drop(columns='_src_ord')


# ---------------------------------------------------------------------------
# 功能推荐（叠加 + 双功能组合）
# ---------------------------------------------------------------------------
def build_recommendations(old):
    main_sources = old['third_src'].value_counts()
    main_sources = main_sources[main_sources >= 200].index.tolist()

    add_recs, pair_recs = [], []

    for src in main_sources:
        grp = old[old['third_src'] == src]
        base = grp['is_paid'].mean()
        n_base = len(grp)
        is_func = src not in NON_FUNC_SOURCES

        all_f = set().union(*grp['use_funcs'])
        for f in all_f:
            if is_func and f == src:
                continue
            sub = grp[grp['use_funcs'].apply(lambda x, ff=f: ff in x)]
            if len(sub) < 80:
                continue
            rate = sub['is_paid'].mean()
            pen = len(sub) / n_base
            lift = rate - base
            if lift >= MIN_LIFT_REC and pen >= MIN_PEN_REC:
                add_recs.append({
                    '入口来源': src,
                    '推荐叠加功能': f,
                    '入口基准付费率': base,
                    '叠加后付费率': rate,
                    'lift': lift,
                    '叠加使用占比': pen,
                    '样本量': len(sub),
                })

        pair_stats = {}
        for _, row in grp.iterrows():
            funcs = sorted(row['use_funcs'])
            pairs = []
            if is_func and src in funcs:
                pairs = [(src, o) for o in funcs if o != src]
            elif not is_func:
                pairs = list(combinations(funcs, 2))
            for p in pairs:
                key = tuple(sorted(p)) if not is_func else p
                if key not in pair_stats:
                    pair_stats[key] = {'n': 0, 'paid': 0}
                pair_stats[key]['n'] += 1
                pair_stats[key]['paid'] += row['is_paid']

        for k, d in pair_stats.items():
            if d['n'] < MIN_PAIR_SAMPLES:
                continue
            rate = d['paid'] / d['n']
            if rate - base >= 0.03:
                pair_recs.append({
                    '入口来源': src,
                    '功能组合': ' + '.join(k),
                    '样本量': d['n'],
                    '付费率': rate,
                    'lift': rate - base,
                    '入口基准': base,
                })

    add_df = pd.DataFrame(add_recs)
    pair_df = pd.DataFrame(pair_recs)
    if not add_df.empty:
        add_df = add_df.sort_values(['入口来源', 'lift'], ascending=[True, False])
    if not pair_df.empty:
        pair_df = pair_df.sort_values(['入口来源', 'lift'], ascending=[True, False])
    return add_df, pair_df


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------
def export_csv(df, path, pct_cols=None):
    out = df.copy()
    if pct_cols:
        for c in pct_cols:
            if c in out.columns:
                out[c] = fmt_pct(out[c])
    out.to_csv(path, index=False, encoding='utf-8-sig')
    print(f'  已输出: {path} ({len(df)} 行)')


def _table2_func_cols(df):
    return [c for c in df.columns if c not in TABLE2_META_COLS]


def export_table2(df, csv_path, xlsx_path):
    """表2：追加功能列列均值行；Excel 中将高于列均值的单元格标红。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    func_cols = _table2_func_cols(df)
    col_means = {c: df[c].mean() for c in func_cols}

    # CSV：数据行 + 列均值行（仅功能列填列均值，其余留空）
    mean_row = {c: np.nan for c in df.columns}
    mean_row['三级来源'] = '列均值'
    for c in func_cols:
        mean_row[c] = col_means[c]
    csv_df = pd.concat([df, pd.DataFrame([mean_row])], ignore_index=True)

    pct_cols = [c for c in csv_df.columns if c not in {'三级来源', '用户数', '付费人数'}]
    export_csv(csv_df, csv_path, pct_cols)

    # Excel：百分比格式 + 高于列均值标红
    wb = Workbook()
    ws = wb.active
    ws.title = '功能使用比例'

    headers = list(df.columns)
    ws.append(headers)

    red_font = Font(color='FF0000', bold=True)
    mean_fill = PatternFill('solid', fgColor='FFF2CC')
    pct_fmt = '0.00%'

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            col_name = headers[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in func_cols and pd.notna(val):
                cell.value = float(val)
                cell.number_format = pct_fmt
                if val > col_means[col_name]:
                    cell.font = red_font
            elif col_name in TABLE2_META_COLS - {'三级来源', '用户数', '付费人数'}:
                if pd.notna(val):
                    cell.value = float(val)
                    cell.number_format = pct_fmt
                else:
                    cell.value = '-'
            else:
                cell.value = val

    # 列均值行
    mean_row_idx = len(df) + 2
    ws.cell(row=mean_row_idx, column=1, value='列均值').font = Font(bold=True)
    for col_idx, col_name in enumerate(headers, start=1):
        if col_name in func_cols:
            cell = ws.cell(row=mean_row_idx, column=col_idx, value=col_means[col_name])
            cell.number_format = pct_fmt
            cell.font = Font(bold=True)
            cell.fill = mean_fill

    wb.save(xlsx_path)
    print(f'  已输出: {xlsx_path} ({len(df)} 行 + 列均值行, 高于均值标红)')


def main():
    base_dir = Path(__file__).parent
    csv_path = base_dir / INPUT_CSV
    print(f'读取: {csv_path}')
    old = load_old_users(csv_path)
    print(f'{USER_TYPE}: {len(old)} 人, 付费率 {old["is_paid"].mean():.2%}\n')

    # 表1
    t1 = build_table1(old)
    export_csv(t1, base_dir / OUTPUT_TABLE1, ['占比', '付费转化率'])

    # 表2（含列均值 + 高于均值标红）
    t2 = build_table2(old)
    export_table2(t2, base_dir / OUTPUT_TABLE2, base_dir / OUTPUT_TABLE2_XLSX)

    # 表3
    t3 = build_table3(old)
    export_csv(
        t3, base_dir / OUTPUT_TABLE3,
        [
            '使用占比', '付费转化率', '入口基准付费率', 'lift',
            '留存率', '入口基准留存率', '留存率lift',
        ],
    )

    # 推荐
    add_df, pair_df = build_recommendations(old)
    if not add_df.empty:
        export_csv(
            add_df, base_dir / OUTPUT_ADD_REC,
            ['入口基准付费率', '叠加后付费率', 'lift', '叠加使用占比'],
        )
    if not pair_df.empty:
        export_csv(
            pair_df, base_dir / OUTPUT_PAIR_REC,
            ['付费率', 'lift', '入口基准'],
        )

    print('\n处理完成。')


if __name__ == '__main__':
    main()
